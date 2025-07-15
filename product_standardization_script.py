import pandas as pd
import json
import re
from openai import OpenAI
import requests
from typing import List, Dict, Any
import logging
import os
import glob

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ProductStandardizer:
    def __init__(self, deepseek_api_key: str, base_url: str = "https://api.deepseek.com"):
        """
        初始化商品标准化器
        
        Args:
            deepseek_api_key: Deepseek API密钥
            base_url: API基础URL
        """
        self.client = OpenAI(
            api_key=deepseek_api_key,
            base_url=base_url
        )
        
        # 标准商品名称列表
        self.standard_products = [
            "四海170g鱼蛋鲜装",
            "四海150g鱼之豆腐鲜装", 
            "四海250g手打香菇贡丸鲜装",
            "四海170g八爪鱼味鱼球鲜装",
            "四海250g手打牛筋丸鲜装",
            "四海250g手打牛肉丸鲜装",
            "四海200g鲜装鱼籽虾饼",
            "四海170g鲜装台湾花枝味鱼丸",
            "四海170g鲜装墨鱼味鱼丸",
            "四海250g墨鱼鱼饼",
            "四海150g鲜装牛肉丸"
        ]
    
    def read_order_data_from_file(self, file_path: str) -> List[str]:
        """
        从order.txt文件中读取订单数据
        
        Args:
            file_path: order.txt文件路径
            
        Returns:
            订单数据列表
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read().strip()
            
            # 按空行分割不同店铺的数据
            orders = []
            current_order = []
            
            for line in content.split('\n'):
                line = line.strip()
                if not line:  # 空行表示一个店铺数据结束
                    if current_order:
                        orders.append('\n'.join(current_order))
                        current_order = []
                else:
                    current_order.append(line)
            
            # 添加最后一个订单
            if current_order:
                orders.append('\n'.join(current_order))
            
            logger.info(f"从 {file_path} 读取了 {len(orders)} 个店铺的订单数据")
            return orders
            
        except Exception as e:
            logger.error(f"读取订单文件失败: {e}")
            return []
        
    def parse_raw_data(self, raw_data: List[str]) -> List[Dict[str, Any]]:
        """
        解析原始数据，提取店铺名称和商品信息
        
        Args:
            raw_data: 原始数据列表
            
        Returns:
            解析后的数据结构
        """
        parsed_data = []
        
        for data in raw_data:
            lines = data.strip().split('\n')
            
            # 查找店铺名称（通常在第一行，以冒号结尾或单独一行）
            shop_name = None
            product_lines = []
            
            for i, line in enumerate(lines):
                line = line.strip()
                if not line:
                    continue
                    
                # 检查是否是店铺名称
                if ':' in line and not re.search(r'\d+件', line):
                    shop_name = line.rstrip(':：')
                    product_lines = lines[i+1:]
                    break
                elif i == 0 and not re.search(r'\d+件', line):
                    shop_name = line
                    product_lines = lines[i+1:]
                    break
            
            # 如果没有找到店铺名称，可能整个数据都是商品信息
            if shop_name is None:
                shop_name = "未知店铺"
                product_lines = lines
            
            # 合并产品信息
            products_text = '\n'.join(product_lines)
            
            parsed_data.append({
                'shopName': shop_name,
                'data': products_text
            })
            
        return parsed_data
    
    def normalize_product_name(self, product_name: str) -> str:
        """
        标准化商品名称，统一格式
        
        Args:
            product_name: 原始商品名称
            
        Returns:
            标准化后的商品名称
        """
        # 去除首尾空格
        name = product_name.strip()
        
        # 统一重量单位：克 -> g
        name = re.sub(r'(\d+)克', r'\1g', name)
        name = re.sub(r'(\d+)G', r'\1g', name)
        
        # 去除多余的空格
        name = re.sub(r'\s+', '', name)
        
        return name
    
    def extract_all_product_variants(self, parsed_data: List[Dict[str, Any]]) -> set:
        """
        提取所有商品变体名称
        
        Args:
            parsed_data: 解析后的数据
            
        Returns:
            所有商品变体的集合
        """
        all_products = set()
        
        for entry in parsed_data:
            lines = entry['data'].strip().split('\n')
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # 解析商品名称
                if ':' in line or '：' in line:
                    product_name = re.split(r'[：:]', line)[0].strip()
                else:
                    # 处理没有冒号的情况，如 "150g鲜装牛肉丸2件"
                    match = re.match(r'(.+?)(\d+件)', line)
                    if match:
                        product_name = match.group(1).strip()
                    else:
                        continue
                
                # 标准化商品名称
                normalized_name = self.normalize_product_name(product_name)
                all_products.add(normalized_name)
                
                # 生成更多变体以便AI更好地理解映射关系
                # 1. 去除重量信息的版本
                name_without_weight = re.sub(r'^\d+g', '', normalized_name)
                if name_without_weight != normalized_name:
                    all_products.add(name_without_weight)
                
                # 2. 去除"鲜装"等修饰词
                name_without_fresh = re.sub(r'鲜装', '', normalized_name)
                if name_without_fresh != normalized_name:
                    all_products.add(name_without_fresh)
                
                # 3. 去除重量和修饰词的版本
                name_clean = re.sub(r'^\d+g', '', name_without_fresh)
                if name_clean != normalized_name and name_clean != name_without_weight:
                    all_products.add(name_clean)
        
        return all_products
    
    def create_product_mapping(self, parsed_data: List[Dict[str, Any]]) -> Dict[str, str]:
        """
        使用Deepseek创建商品名称映射
        
        Args:
            parsed_data: 解析后的数据
            
        Returns:
            商品名称映射字典
        """
        # 提取所有商品变体
        all_products = self.extract_all_product_variants(parsed_data)
        
        # 构建更详细的提示词
        prompt = f"""
请帮我将以下商品名称（包括各种变体）映射到标准的商品全称。

标准商品全称列表：
{json.dumps(self.standard_products, ensure_ascii=False, indent=2)}

需要映射的商品名称（包括简写和变体）：
{json.dumps(list(all_products), ensure_ascii=False, indent=2)}

映射规则和示例：
1. 优先根据重量信息精确匹配：
   - "150g鲜装牛肉丸" → "四海150g鲜装牛肉丸"
   - "170g鱼蛋鲜装" → "四海170g鱼蛋鲜装"
   - "250g手打牛筋丸" → "四海250g手打牛筋丸鲜装"

2. 关键词匹配（无重量信息时选择最常见规格）：
   - "牛肉丸" → "四海150g鲜装牛肉丸"（默认规格）
   - "鱼蛋鲜装" → "四海170g鱼蛋鲜装"
   - "手打牛肉丸" → "四海250g手打牛肉丸鲜装"

3. 同义词处理：
   - "香菇贡丸" = "手打香菇贡丸" → "四海250g手打香菇贡丸鲜装"
   - "台湾花枝味丸" = "台湾花枝味鱼丸" → "四海170g鲜装台湾花枝味鱼丸"
   - "八爪鱼味鱼球" → "四海170g八爪鱼味鱼球鲜装"

4. 变体处理：
   - "手打香茹丸"（错别字"茹"）→ "四海250g手打香菇贡丸鲜装"
   - "克" = "g"：统一为g

请返回JSON格式的映射关系，每个输入名称都必须有对应的标准全称：
{{
    "鱼蛋鲜装": "四海170g鱼蛋鲜装",
    "170g鱼蛋鲜装": "四海170g鱼蛋鲜装",
    "牛肉丸": "四海150g鲜装牛肉丸"
}}

只返回JSON格式，不要其他说明文字。
"""
        
        try:
            response = self.client.chat.completions.create(
                model="deepseek-chat",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1
            )
            
            # 解析响应
            response_text = response.choices[0].message.content
            logger.debug(f"Deepseek原始响应: {response_text}")
            
            # 提取JSON部分
            json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
            if json_match:
                mapping = json.loads(json_match.group())
                logger.info(f"成功创建商品映射: {len(mapping)} 个商品变体")
                return mapping
            else:
                logger.error("无法从响应中提取JSON")
                return {}
                
        except Exception as e:
            logger.error(f"调用Deepseek API失败: {e}")
            return {}
    
    def standardize_data(self, parsed_data: List[Dict[str, Any]], product_mapping: Dict[str, str]) -> List[Dict[str, Any]]:
        """
        标准化数据
        
        Args:
            parsed_data: 解析后的数据
            product_mapping: 商品名称映射
            
        Returns:
            标准化后的数据
        """
        standardized_data = []
        
        for entry in parsed_data:
            shop_name = entry['shopName']
            lines = entry['data'].strip().split('\n')
            
            shop_products = {}
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                    
                # 解析商品和数量
                product_name = None
                quantity = None
                
                if ':' in line or '：' in line:
                    # 标准格式：商品名：数量
                    parts = re.split(r'[：:]', line)
                    if len(parts) >= 2:
                        product_name = parts[0].strip()
                        quantity_str = parts[1].strip()
                        
                        # 提取数量
                        quantity_match = re.search(r'(\d+)件', quantity_str)
                        if quantity_match:
                            quantity = int(quantity_match.group(1))
                else:
                    # 无冒号格式：150g鲜装牛肉丸2件
                    match = re.match(r'(.+?)(\d+)件', line)
                    if match:
                        product_name = match.group(1).strip()
                        quantity = int(match.group(2))
                
                if product_name and quantity:
                    # 标准化商品名称
                    normalized_name = self.normalize_product_name(product_name)
                    
                    # 查找标准名称（先尝试完整匹配，再尝试去除重量的匹配）
                    standard_name = product_mapping.get(normalized_name)
                    if not standard_name:
                        # 尝试去除重量信息的匹配
                        name_without_weight = re.sub(r'^\d+g', '', normalized_name)
                        standard_name = product_mapping.get(name_without_weight)
                    
                    if not standard_name:
                        standard_name = normalized_name
                        logger.warning(f"未找到 '{normalized_name}' 的映射，使用原名称")
                    
                    shop_products[standard_name] = quantity
            
            standardized_data.append({
                'shopName': shop_name,
                'products': shop_products
            })
            
        return standardized_data
    
    def update_excel_file(self, file_path: str, standardized_data: List[Dict[str, Any]]):
        """
        直接更新Excel文件，保持原有格式和样式
        
        Args:
            file_path: Excel文件路径
            standardized_data: 标准化后的数据
        """
        try:
            from openpyxl import load_workbook
            
            # 使用openpyxl加载工作簿以保持格式
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            
            # 获取表头信息（第二行包含店铺名称）
            header_row = []
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=2, column=col).value
                header_row.append(cell_value)
            
            # 创建店铺名称到列索引的映射
            shop_column_mapping = {}
            for i, col_name in enumerate(header_row):
                if col_name and col_name not in ['序号', '商品编码', '商品名称', '规格', '入库价', '售价', '前台毛利', '供应商编码', '供应商名称']:
                    shop_column_mapping[col_name] = i + 1  # openpyxl使用1基索引
            
            logger.info(f"找到店铺列: {list(shop_column_mapping.keys())}")
            
            # 更新数据
            for entry in standardized_data:
                shop_name = entry['shopName']
                products = entry['products']
                
                # 清理店铺名称（去除冒号等符号）
                clean_shop_name = shop_name.rstrip('：:').strip()
                
                # 查找对应的列索引
                target_column_index = None
                for col_name, col_index in shop_column_mapping.items():
                    if clean_shop_name == col_name or clean_shop_name in col_name or col_name in clean_shop_name:
                        target_column_index = col_index
                        break
                
                # 如果没有找到精确匹配，尝试部分匹配
                if target_column_index is None:
                    for col_name, col_index in shop_column_mapping.items():
                        # 处理特殊情况
                        if ('五江' in clean_shop_name and '五江' in col_name) or \
                           ('金海' in clean_shop_name and '金海' in col_name) or \
                           ('洋湖' in clean_shop_name and '洋湖' in col_name) or \
                           ('砂之船' in clean_shop_name and '砂之船' in col_name) or \
                           ('邵阳' in clean_shop_name and '邵阳' in col_name) or \
                           ('岳阳' in clean_shop_name and '岳阳' in col_name):
                            target_column_index = col_index
                            break
                
                if target_column_index is None:
                    logger.warning(f"未找到店铺 '{shop_name}' (清理后: '{clean_shop_name}') 对应的列")
                    continue
                
                # 更新商品数量
                for product_name, quantity in products.items():
                    # 查找商品行（从第3行开始）
                    product_found = False
                    for row in range(3, worksheet.max_row + 1):
                        cell_value = worksheet.cell(row=row, column=3).value  # 商品名称在第3列
                        if cell_value and product_name in str(cell_value):
                            # 只更新数值，不改变格式
                            worksheet.cell(row=row, column=target_column_index).value = quantity
                            logger.info(f"更新 {clean_shop_name} - {product_name}: {quantity}件")
                            product_found = True
                            break
                    
                    if not product_found:
                        logger.warning(f"未找到商品: {product_name}")
            
            # 保存工作簿，保持原有格式
            workbook.save(file_path)
            logger.info(f"Excel文件已更新: {file_path}")
            
        except Exception as e:
            logger.error(f"更新Excel文件失败: {e}")
    
    def find_excel_file(self, directory: str = ".") -> str:
        """
        自动检测目录中的xlsx文件
        
        Args:
            directory: 要搜索的目录，默认为当前目录
            
        Returns:
            找到的Excel文件路径，如果没找到则返回None
        """
        # 搜索xlsx文件
        xlsx_files = glob.glob(os.path.join(directory, "*.xlsx"))
        
        if not xlsx_files:
            logger.error(f"在目录 {directory} 中未找到任何xlsx文件")
            return None
        elif len(xlsx_files) == 1:
            logger.info(f"找到Excel文件: {xlsx_files[0]}")
            return xlsx_files[0]
        else:
            # 如果有多个xlsx文件，优先选择包含"订单"、"模板"等关键词的文件
            priority_keywords = ["订单", "模板", "template", "order"]
            for keyword in priority_keywords:
                for file_path in xlsx_files:
                    if keyword in os.path.basename(file_path).lower():
                        logger.info(f"找到优先Excel文件: {file_path}")
                        return file_path
            
            # 如果没有匹配关键词的，选择第一个
            logger.info(f"找到多个Excel文件，选择第一个: {xlsx_files[0]}")
            logger.info(f"所有找到的文件: {xlsx_files}")
            return xlsx_files[0]
    
    def process_data_from_file(self, order_file_path: str = "order.txt", excel_file_path: str = None):
        """
        从order.txt文件处理数据并更新Excel文件
        
        Args:
            order_file_path: order.txt文件路径，默认为"order.txt"
            excel_file_path: Excel文件路径，如果为None则自动检测
        """
        logger.info("开始处理数据...")
        
        # 如果没有指定Excel文件路径，自动检测
        if excel_file_path is None:
            excel_file_path = self.find_excel_file()
            if excel_file_path is None:
                logger.error("无法找到Excel文件，处理终止")
                return
        
        # 1. 从文件读取原始数据
        raw_data = self.read_order_data_from_file(order_file_path)
        if not raw_data:
            logger.error("没有读取到订单数据")
            return
        
        # 2. 解析原始数据
        parsed_data = self.parse_raw_data(raw_data)
        logger.info(f"解析了 {len(parsed_data)} 个店铺的数据")
        
        # 3. 创建商品映射
        product_mapping = self.create_product_mapping(parsed_data)
        # 打印 product_mapping 数据
        logger.info(f"商品映射: {json.dumps(product_mapping, ensure_ascii=False, indent=2)}")
        
        # 4. 标准化数据
        standardized_data = self.standardize_data(parsed_data, product_mapping)
        # 打印 standardized_data 数据
        logger.info(f"标准化数据: {json.dumps(standardized_data, ensure_ascii=False, indent=2)}")
        
        # 5. 直接更新Excel文件
        self.update_excel_file(excel_file_path, standardized_data)
        
        logger.info("数据处理完成！")

# 使用示例
if __name__ == "__main__":
    # 配置Deepseek API
    API_KEY = "sk-xxxxxxxxxxxxxxxxxx"  # 替换为你的API密钥
    
    # 创建处理器实例
    processor = ProductStandardizer(API_KEY)
    
    # 处理数据 - 现在会自动检测xlsx文件
    processor.process_data_from_file(
        order_file_path="order.txt"  # order.txt文件路径
        # excel_file_path 参数已移除，脚本会自动检测根目录中的xlsx文件
    )
